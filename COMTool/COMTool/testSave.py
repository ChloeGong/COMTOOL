#窗口：name+path 
# button:cancel+save 
#窗口一定要有 _init_ init initEvent closeEvent
from PyQt5.QtCore import pyqtSignal,Qt,QRect
from PyQt5.QtWidgets import (QApplication, QWidget,QToolTip,QPushButton,QMessageBox,QDesktopWidget,QMainWindow,
                             QVBoxLayout,QHBoxLayout,QGridLayout,QTextEdit,QLabel,QRadioButton,QCheckBox,
                             QLineEdit,QGroupBox,QSplitter,QDialog)
from Combobox import ComboBox
import xlrd
import xlwt
# workbook相关
from openpyxl.workbook import Workbook
# ExcelWriter，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter
# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook

class Ui_TestSave(QWidget):#窗口设计  
    def initwindow(self):
        #设定窗口属性
        self.setObjectName("Save File")
        self.resize(500,400)#窗口的大小
        #定义控件
        self.label = QLabel()
        self.saveButton = QPushButton("Save")
        self.cancelButton = QPushButton("Cancel")

        #定义布局
        self.mainLayout = QVBoxLayout()
        self.frameLayout = QHBoxLayout()
        self.buttonLayout = QHBoxLayout()



        self.savePathCombobox = ComboBox()#下拉列表
        self.savePathCombobox.addItem("Desktop")
        self.savePathCombobox.addItem("C:/")
        self.savePathCombobox.addItem("D:/")
        self.savePathCombobox.addItem("E:/")
        



        #将控件加入布局
        self.frameLayout.addWidget(self.label)
        self.frameLayout.addWidget(self.savePathCombobox)
        self.buttonLayout.addWidget(self.saveButton)
        self.buttonLayout.addWidget(self.cancelButton)
        self.mainLayout.addLayout(self.frameLayout)
        self.mainLayout.addLayout(self.buttonLayout)
        
        self.mainLayout.setStretch(0,1)
        self.mainLayout.setStretch(1,1)

        #将布局加入窗体
        self.setLayout(self.mainLayout)



        #设定控件属性
        self.label.setText("Save Path")
        self.cancelButton.clicked.connect(self.closeEvent)
        
        
       
    def __init__(self,parent = None):
        super(Ui_TestSave,self).__init__(parent)
        self.initwindow()
        self.show()
       # self. head_row_labels = [u'StudentId',u'Name',u'Number',u'SubjectId',u'Subject']

    def write_to_excel_with_openpyxl(self,records,head_row,save_excel_name="save.xlsx"):
        wb = Workbook()
       # ew = ExcelWriter(workbook=wb)
        # 设置文件输出路径与名称
        dest_filename = save_excel_name.decode('utf-8')
        # 第一个sheet是ws
        ws = wb.worksheets[0]
        # 设置ws的名称
        ws.title = "range names"
        # 写第一行，标题行
        for h_x in range(1,len(head_row)+1):
            h_col=get_column_letter(h_x)
            #print h_col
            ws.cell('%s%s' % (h_col, 1)).value = '%s' % (head_row[h_x-1])
        # 写第二行及其以后的那些行
        i = 2
        for record in records:
            record_list=str(record).strip().split("\t")
            for x in range(1,len(record_list)+1):
                col = get_column_letter(x)
                ws.cell('%s%s' % (col, i)).value = '%s' % (record_list[x-1].decode('utf-8'))
            i += 1
        ew.save(filename=dest_filename)
    def write_to_excel(self, dataset,save_excel_name,head_row):
        f = xlwt.Workbook()  # 创建工作簿
        # 创建第一个sheet:
        # sheet1
        count=1
        sheet1 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
        # 首行标题：
        for p in  range(len(head_row)):
                sheet1.write(0,p,head_row[p],self.set_style('Times New Roman',250,True))
        default=self.set_style('Times New Roman',200,False)   # define style out the loop will work
        for line in dataset:
            row_list=str(line).strip("\n").split("\t")
            for pp in  range(len(str(line).strip("\n").split("\t"))):
                sheet1.write(count,pp,row_list[pp].decode('utf-8'),default)
            count+=1
        f.save(save_excel_name)  # 保存文件
    def run_main_save_to_excel_with_openpyxl(self):
        #dataset_list = self.read_from_file("test_excel.txt")
        head_row_label = self.head_row_labels
        save_name = "test_openpyxl.xlsx"
        #self.write_to_excel_with_openpyxl(dataset_list,head_row_label,save_name)
    def run_main_save_to_excel_with_xlwt(self):
        #dataset_list=self.read_from_file("test_excel.txt")
        head_row_label=self.head_row_labels
        save_name="test_xlwt.xls"
        #self.write_to_excel_with_openpyxl(dataset_list,head_row_label,save_name)  

    def closeEvent(self,event):
        reply = QMessageBox.question(self,'本程序',"是否要退出保存？",QMessageBox.Yes|QMessageBox.No,QMessageBox.No) # self,"title",“information”
        if reply == QMessageBox.Yes:
           self.closed.emit() 
           event.accept()
        else:
           event.ignore()

            