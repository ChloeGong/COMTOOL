#coding:utf-8
import xlrd
import xlwt
# workbook相关
from openpyxl.workbook import Workbook
# ExcelWriter，封装了很强大的excel写的功能
from openpyxl.writer.excel import ExcelWriter
# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
class HandleExcel():
    '''Excel相关操作类'''
    def __init__(self):
        self. head_row_labels = [u'StudentId',u'Name',u'Number',u'SubjectId',u'Subject']
    def write_to_excel_with_openpyxl(self,records,head_row,save_excel_name="save.xlsx"):
        wb = Workbook()
       # ew = ExcelWriter(workbook=wb)
        # 设置文件输出路径与名称
        dest_filename = save_excel_name.decode('utf-8')
        # 第一个sheet是ws
        ws = wb.worksheets[0]
        # 设置ws的名称
        ws.title = "range names"
        # 写第一行，标题行
        for h_x in range(1,len(head_row)+1):
            h_col=get_column_letter(h_x)
            #print h_col
            ws.cell('%s%s' % (h_col, 1)).value = '%s' % (head_row[h_x-1])
        # 写第二行及其以后的那些行
        i = 2
        for record in records:
            record_list=str(record).strip().split("\t")
            for x in range(1,len(record_list)+1):
                col = get_column_letter(x)
                ws.cell('%s%s' % (col, i)).value = '%s' % (record_list[x-1].decode('utf-8'))
            i += 1
        ew.save(filename=dest_filename)
    def set_style(self,name,height,bold=False):
        style = xlwt.XFStyle() # 初始化样式
        font = xlwt.Font() # 为样式创建字体
        font.name = name  # 'Times New Roman'
        font.bold = bold
        font.color_index = 4
        font.height = height
        borders= xlwt.Borders()
        borders.left= 6
        borders.right= 6
        borders.top= 6
        borders.bottom= 6
        style.font = font
        style.borders = borders
        return style
    def write_to_excel(self, dataset,save_excel_name,head_row):
        f = xlwt.Workbook()  # 创建工作簿
        # 创建第一个sheet:
        # sheet1
        count=1
        sheet1 = f.add_sheet(u'sheet1', cell_overwrite_ok=True)  # 创建sheet
        # 首行标题：
        for p in  range(len(head_row)):
                sheet1.write(0,p,head_row[p],self.set_style('Times New Roman',250,True))
        default=self.set_style('Times New Roman',200,False)   # define style out the loop will work
        for line in dataset:
            row_list=str(line).strip("\n").split("\t")
            for pp in  range(len(str(line).strip("\n").split("\t"))):
                sheet1.write(count,pp,row_list[pp].decode('utf-8'),default)
            count+=1
        f.save(save_excel_name)  # 保存文件
    def run_main_save_to_excel_with_openpyxl(self):
        #dataset_list = self.read_from_file("test_excel.txt")
        head_row_label = self.head_row_labels
        save_name = "test_openpyxl.xlsx"
        #self.write_to_excel_with_openpyxl(dataset_list,head_row_label,save_name)
    def run_main_save_to_excel_with_xlwt(self):
        #dataset_list=self.read_from_file("test_excel.txt")
        head_row_label=self.head_row_labels
        save_name="test_xlwt.xls"
        #self.write_to_excel_with_openpyxl(dataset_list,head_row_label,save_name)
if __name__ == '__main__':
    obj_handle_excel=HandleExcel()
    obj_handle_excel.run_main_save_to_excel_with_openpyxl()
    obj_handle_excel.run_main_save_to_excel_with_xlwt()
    